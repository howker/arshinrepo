"""Тесты на регрессии, пойманные в бою.

Каждый тест здесь соответствует реальному багу, который дошёл до продакшена.
Ответы Аршина зафиксированы — сеть не используется.
"""
from __future__ import annotations

from datetime import date

import pytest

from app.clients.arshin import ArshinSearchResult
from app.excel.detector import detect_layout, LayoutNotRecognizedError
from app.models.enums import CheckResultClass, JobItemStatus
from app.services.comparator import compare_device
from app.services.matcher import (
    MatchResult,
    ensure_date,
    model_match_score,
    select_best_match,
    type_confirmation_score,
)


# --- Реальные ответы Аршина, снятые с боевых запросов ---

ARSHIN_ZNOL_6590 = [
    # Настоящий прибор метролога (Астрахань)
    {"vri_id": "1-256311832", "mi_number": "6590", "mi_mitnumber": "3344-08",
     "mi_type": None, "mi_modification": "модификация ЗНОЛ.06",
     "verification_date": "2023-06-12T00:00:00Z", "valid_date": "2031-06-11T00:00:00Z",
     "applicability": True, "org_title": 'ФБУ "АСТРАХАНСКИЙ ЦСМ"',
     "card_url": "https://fgis.gost.ru/fundmetrology/cm/results/1-256311832"},
    # Чужой ЗНОЛ с тем же заводским номером (Вологда)
    {"vri_id": "2-169084007", "mi_number": "6590", "mi_mitnumber": "3344-08",
     "mi_type": None, "mi_modification": "ЗНОЛ.06",
     "verification_date": "2019-09-07T12:00:00Z", "valid_date": "2027-09-06T00:00:00Z",
     "applicability": True, "org_title": 'ФБУ "ВОЛОГОДСКИЙ ЦСМ"',
     "card_url": "https://fgis.gost.ru/fundmetrology/cm/results/2-169084007"},
]

ARSHIN_EVROALFA = [
    # Тип лежит в mi_type, а не в mi_modification
    {"vri_id": "1-25798593", "mi_number": "01111479", "mi_mitnumber": "16666-97",
     "mi_type": "ЕвроАЛЬФА", "mi_modification": "EA05",
     "verification_date": "2020-11-20T00:00:00Z", "valid_date": "2028-11-19T00:00:00Z",
     "applicability": True, "org_title": 'ФБУ "АСТРАХАНСКИЙ ЦСМ"',
     "card_url": "https://fgis.gost.ru/fundmetrology/cm/results/1-25798593"},
]

ARSHIN_FOREIGN_TSHL = [
    # ТШЛ — не ТЛШ. Перестановка букв = другой прибор.
    {"vri_id": "1-7948626", "mi_number": "2301", "mi_mitnumber": "11077-03",
     "mi_type": None, "mi_modification": "ТШЛ-0,66-2 У2",
     "verification_date": "2020-09-24T00:00:00Z", "valid_date": "2028-09-23T00:00:00Z",
     "applicability": True, "org_title": 'ООО "ТПК "ЭНЕРГОУЧЕТ"',
     "card_url": "https://fgis.gost.ru/fundmetrology/cm/results/1-7948626"},
]

ASTRAKHAN = {"sub_company": "ООО Газпром добыча Астрахань", "object_name": "ГПП 1"}


def _result(records):
    return ArshinSearchResult(
        records=records,
        result_class=CheckResultClass.SUCCESS_WITH_MATCH,
        candidates_count=len(records),
    )


class TestIdentification:
    """Идентификация прибора: никогда не подставлять чужой."""

    def test_region_picks_the_right_duplicate(self):
        """Два ЗНОЛ с одним номером — выбираем по региону поверителя."""
        m = select_best_match(
            "6590", "ЗНОЛ0610УЗ", _result(ARSHIN_ZNOL_6590),
            item_type_raw="ЗНОЛ-0.6-10 УЗ", owner_context=ASTRAKHAN,
        )
        assert m.result_class == CheckResultClass.SUCCESS_WITH_MATCH
        assert m.selected["vri_id"] == "1-256311832"

    def test_typos_do_not_break_identification(self):
        """Метролог пишет тип как попало — прибор всё равно находится."""
        for typed in ["ЗНОЛ-0.6-10 УЗ", "ЗНОЛ-0.6-10-УЗ", "ЗНОЛ 0,6-10 УЗ", "знол.06"]:
            m = select_best_match(
                "6590", "ЗНОЛ", _result(ARSHIN_ZNOL_6590),
                item_type_raw=typed, owner_context=ASTRAKHAN,
            )
            assert m.selected["vri_id"] == "1-256311832", f"сломалось на {typed!r}"

    def test_foreign_device_never_substituted(self):
        """ТШЛ вместо ТЛШ — не выдаём за найденный прибор."""
        m = select_best_match(
            "2301", "ТЛШ10", _result(ARSHIN_FOREIGN_TSHL), item_type_raw="ТЛШ-10",
        )
        assert m.result_class == CheckResultClass.AMBIGUOUS_MULTIPLE_MATCHES

    def test_type_may_live_in_mitype_field(self):
        """ЕвроАЛЬФА: тип в mi_type, модификация EA05 — не должно быть расхождением."""
        m = select_best_match(
            "01111479", "ЕВРОАЛЬФА", _result(ARSHIN_EVROALFA), item_type_raw="ЕвроАльфа",
        )
        assert m.result_class == CheckResultClass.SUCCESS_WITH_MATCH


class TestDates:
    """Даты Аршина приходят в ISO-8601 — сравнение молча не пропускалось."""

    def test_iso_dates_parsed(self):
        assert ensure_date("2023-06-12T00:00:00Z") == date(2023, 6, 12)
        assert ensure_date("2019-09-07T12:00:00Z") == date(2019, 9, 7)
        assert ensure_date("INVALID") is None

    def test_date_mismatch_is_detected(self):
        """Главный баг: расхождение даты обязано находиться."""
        m = select_best_match(
            "6590", "ЗНОЛ0610УЗ", _result(ARSHIN_ZNOL_6590),
            item_type_raw="ЗНОЛ-0.6-10 УЗ", owner_context=ASTRAKHAN,
        )
        device = {
            "type_raw": "ЗНОЛ-0.6-10 УЗ",
            "type_norm": "ЗНОЛ0610УЗ",
            "serial_norm": "6590",
            "verification_date_norm": date(2023, 6, 12),
            "next_date_norm": date(2031, 6, 12),   # в Аршине 11-е — расхождение!
            "link_raw": None,
            "link_vri": None,
            "cell_refs": {"type": "W11", "serial": "Y11",
                          "verification_date": "AB11", "next_date": "AC11", "link": "AD11"},
        }
        r = compare_device(device, m)
        codes = [i.code for i in r.issues]
        assert "NEXT_VERIFICATION_DATE_MISMATCH" in codes, codes
        assert r.status == JobItemStatus.MISMATCH


class TestModelComparison:
    """Сверка модели строже, чем идентификация семейства."""

    def test_family_root_is_not_enough_for_comparison(self):
        assert model_match_score("ТЛШ-10", "ТЛШ-20") < 85.0     # разные приборы
        assert model_match_score("ЗНОЛ-0.6-10 УЗ", "модификация ЗНОЛ.06") >= 85.0
        assert type_confirmation_score("ТЛШ-10", "ТШЛ-0,66-2 У2") < 75.0


class TestSheetDetection:
    """Раскладка распознаётся по структуре, а не по имени листа."""

    def _sheet(self, header_row: int, sheet_title: str):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title
        # Группы приборов
        ws.cell(row=header_row - 2, column=8).value = "Счетчик электрической энергии"
        ws.cell(row=header_row - 2, column=15).value = "Трансформатор тока"
        # Заголовки полей
        ws.cell(row=header_row, column=8).value = "Тип ПУ"
        ws.cell(row=header_row, column=9).value = "Зав. номер"
        ws.cell(row=header_row, column=10).value = "Дата поверки"
        ws.cell(row=header_row, column=15).value = "Тип ТТ"
        ws.cell(row=header_row, column=16).value = "Зав. номер"
        ws.cell(row=header_row, column=17).value = "Дата поверки"
        # Данные
        ws.cell(row=header_row + 1, column=8).value = "Меркурий 234"
        ws.cell(row=header_row + 1, column=9).value = "40724005"
        return ws

    @pytest.mark.parametrize("title", ["Отчет", "Лист1", "Прил.1.1 (Сч,ТТ,ТН)", "ведомость"])
    def test_any_sheet_name(self, title):
        """Имя листа не должно влиять на распознавание."""
        layout = detect_layout(self._sheet(8, title))
        assert layout is not None
        assert "si" in layout.groups

    @pytest.mark.parametrize("header_row", [8, 14, 20])
    def test_any_header_position(self, header_row):
        """Метролог мог вставить сверху подписи — шапка съезжает."""
        layout = detect_layout(self._sheet(header_row, "Отчет"))
        assert layout is not None
        assert layout.header_row == header_row

    def test_not_a_sheet_of_devices(self):
        """Посторонний лист не должен опознаваться как ведомость."""
        from openpyxl import Workbook
        ws = Workbook().active
        ws["A1"] = "Смета"
        ws["A2"] = "Итого"
        assert detect_layout(ws) is None


class TestProgressTracking:
    """Прогресс не должен затираться проверкой отмены.

    Баг: воркер делал db.refresh(job) на каждой итерации, чтобы узнать,
    не отменили ли задачу. refresh() перечитывал объект целиком и сбрасывал
    несохранённые поля — статус откатывался в «в очереди», а счётчик
    обработанных приборов обнулялся, хотя работа шла.
    """

    def test_refresh_wipes_pending_changes(self):
        """Демонстрация механики бага на SQLAlchemy (без БД приложения)."""
        from sqlalchemy import Column, Integer, String, create_engine
        from sqlalchemy.orm import declarative_base, Session

        Base = declarative_base()

        class Row(Base):
            __tablename__ = "t"
            id = Column(Integer, primary_key=True)
            status = Column(String)
            processed = Column(Integer, default=0)

        engine = create_engine("sqlite://")
        Base.metadata.create_all(engine)

        with Session(engine) as s:
            s.add(Row(id=1, status="queued", processed=0))
            s.commit()

            row = s.get(Row, 1)
            row.status = "processing"     # ещё не закоммичено
            row.processed = 42

            s.refresh(row)                # <-- так делал воркер

            assert row.status == "queued", "refresh обязан затереть — это и был баг"
            assert row.processed == 0

    def test_scalar_query_preserves_pending_changes(self):
        """Правильный способ: точечный запрос не трогает объект в памяти."""
        from sqlalchemy import Column, Integer, String, create_engine, select
        from sqlalchemy.orm import declarative_base, Session

        Base = declarative_base()

        class Row(Base):
            __tablename__ = "t2"
            id = Column(Integer, primary_key=True)
            status = Column(String)
            processed = Column(Integer, default=0)

        engine = create_engine("sqlite://")
        Base.metadata.create_all(engine)

        with Session(engine) as s:
            s.add(Row(id=1, status="queued", processed=0))
            s.commit()

            row = s.get(Row, 1)
            row.status = "processing"
            row.processed = 42

            # Так делает воркер: no_autoflush, иначе SELECT сначала запишет
            # наши несохранённые изменения в БД и проверка станет нечестной.
            with s.no_autoflush:
                current = s.execute(select(Row.status).where(Row.id == 1)).scalar()

            assert current == "queued"          # в БД пока старое — отмены не было
            assert row.status == "processing"   # прогресс в памяти НЕ затёрт
            assert row.processed == 42


def test_worker_does_not_use_refresh():
    """Страховка: db.refresh(job) не должен вернуться в код воркера."""
    import re
    from pathlib import Path

    src = Path("app/workers/tasks.py").read_text(encoding="utf-8")
    code = "\n".join(
        line for line in src.split("\n")
        if not line.strip().startswith("#")
    )
    assert not re.search(r"\bdb\.refresh\(job\)", code), (
        "db.refresh(job) затирает несохранённый прогресс — "
        "проверяйте отмену точечным запросом статуса"
    )

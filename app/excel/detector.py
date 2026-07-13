"""Структурное распознавание ведомости СИ.

Метролог волен назвать лист как угодно («Отчет», «Лист1», «Прил.1.1»),
вставить сверху лишние строки с подписями и сдвинуть колонки. Поэтому
привязываться к имени листа и номерам колонок нельзя — распознаём документ
по СТРУКТУРЕ: находим строку заголовков и определяем колонки по их названиям.
"""
from __future__ import annotations

import logging
import re

from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

MAX_SCAN_ROWS = 40
MAX_SCAN_COLS = 60


def norm_header(value) -> str:
    if value is None:
        return ""
    s = re.sub(r"\s+", " ", str(value)).strip().lower()
    return s.replace("ё", "е")


# Заголовки полей внутри группы приборов
FIELD_ALIASES: dict[str, list[str]] = {
    "type": ["тип пу", "тип тт", "тип тн", "тип си", "тип счетчика", "тип"],
    "transformation_ratio": [
        "коэф. трансформации", "коэф трансформации", "коэффициент трансформации",
    ],
    "serial": ["зав. номер", "зав номер", "зав.номер", "заводской номер", "заводской №"],
    "accuracy_class": ["класс точности"],
    "manufacture_year": ["год выпуска"],
    "verification_date": ["дата поверки"],
    "next_verification_date": [
        "дата след. поверки", "дата след поверки", "дата следующей поверки",
        "дата след.поверки",
    ],
    "arshin_link": [
        "наличие св-ва о поверке (паспорта)", "наличие св-ва о поверке",
        "св-во о поверке", "свидетельство о поверке", "ссылка на аршин", "ссылка",
    ],
}

# Заголовки групп приборов (объединённая шапка над полями)
GROUP_ALIASES: dict[str, list[str]] = {
    "si": ["счетчик электрической энергии", "счетчики электрической энергии", "счетчик"],
    "ct": ["трансформатор тока", "трансформаторы тока"],
    "vt": ["трансформатор напряжения", "трансформаторы напряжения"],
}

CONTEXT_ALIASES: dict[str, list[str]] = {
    "n_pp": ["№ п/п", "n п/п", "№п/п"],
    "sub_company": ["дочернее общество"],
    "object_name": ["наименование объекта"],
    "inventory_no": ["инвентарный номер"],
    "connection_point": ["наименование точки присоединения", "наименование точки"],
    "voltage_level": ["уровень напряжения в точке учета", "уровень напряжения"],
    "metering_type": ["тип учета"],
}


def _match(value, aliases: dict[str, list[str]]) -> str | None:
    v = norm_header(value)
    if not v:
        return None
    for key, variants in aliases.items():
        for a in variants:
            if v == a or v.startswith(a):
                return key
    return None


class SheetLayout:
    """Распознанная раскладка ведомости."""

    def __init__(
        self,
        sheet: Worksheet,
        header_row: int,
        data_start_row: int,
        groups: dict[str, dict[str, int]],
        context: dict[str, int],
        max_column: int,
    ):
        self.sheet = sheet
        self.header_row = header_row
        self.data_start_row = data_start_row
        self.groups = groups        # {"si": {"type": 7, "serial": 8, ...}, ...}
        self.context = context      # {"sub_company": 2, ...}
        self.max_column = max_column


class LayoutNotRecognizedError(ValueError):
    pass


def _find_header_row(ws: Worksheet) -> tuple[int, dict[str, list[int]]] | None:
    """Строка заголовков — та, где есть и «Зав. номер», и «Дата поверки»."""
    for r in range(1, min(ws.max_row, MAX_SCAN_ROWS) + 1):
        found: dict[str, list[int]] = {}
        for c in range(1, min(ws.max_column, MAX_SCAN_COLS) + 1):
            key = _match(ws.cell(row=r, column=c).value, FIELD_ALIASES)
            if key:
                found.setdefault(key, []).append(c)
        if "serial" in found and "verification_date" in found:
            return r, found
    return None


def _first_data_row(ws: Worksheet, header_row: int, serial_cols: list[int]) -> int:
    """Первая строка с настоящими данными.

    Сразу под шапкой может лежать мусор: строка нумерации колонок,
    формулы вида #NAME?, пустые строки. Ищем первую строку, где хотя бы
    в одной колонке серийника есть осмысленное значение.
    """
    junk = ("#name?", "#имя?", "#ref!", "#value!", "#знач!")
    for r in range(header_row + 1, min(ws.max_row, header_row + 12) + 1):
        for c in serial_cols:
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if not s or s in junk or s in ("-", "—"):
                continue
            # Строка нумерации колонок («1», «2», «3»…) — не данные
            row_vals = [ws.cell(row=r, column=cc).value for cc in range(1, 8)]
            digits = sum(1 for x in row_vals if str(x).strip().isdigit())
            if digits >= 5:
                break
            return r
    return header_row + 1


def detect_layout(ws: Worksheet) -> SheetLayout | None:
    """Распознаёт раскладку на листе. None — если это не ведомость СИ."""
    hit = _find_header_row(ws)
    if not hit:
        return None
    header_row, fields = hit

    # Границы групп приборов — из объединённой шапки над строкой заголовков
    group_starts: list[tuple[int, str]] = []
    for r in range(max(1, header_row - 4), header_row):
        for c in range(1, min(ws.max_column, MAX_SCAN_COLS) + 1):
            g = _match(ws.cell(row=r, column=c).value, GROUP_ALIASES)
            if g and all(g != x[1] for x in group_starts):
                group_starts.append((c, g))
    if not group_starts:
        return None
    group_starts.sort()

    # Каждое поле относится к той группе, в чей диапазон колонок попадает
    bounds: list[tuple[str, int, int]] = []
    for i, (col, kind) in enumerate(group_starts):
        end = group_starts[i + 1][0] - 1 if i + 1 < len(group_starts) else ws.max_column
        bounds.append((kind, col, end))

    groups: dict[str, dict[str, int]] = {kind: {} for kind, _, _ in bounds}
    for field, cols in fields.items():
        for c in cols:
            for kind, start, end in bounds:
                if start <= c <= end:
                    groups[kind].setdefault(field, c)
                    break

    # Группа считается пригодной, если есть чем идентифицировать прибор
    groups = {
        k: v for k, v in groups.items()
        if "serial" in v and "type" in v
    }
    if not groups:
        return None

    context: dict[str, int] = {}
    for r in range(max(1, header_row - 5), header_row + 1):
        for c in range(1, min(ws.max_column, MAX_SCAN_COLS) + 1):
            key = _match(ws.cell(row=r, column=c).value, CONTEXT_ALIASES)
            if key:
                context.setdefault(key, c)

    serial_cols = [g["serial"] for g in groups.values()]
    data_start = _first_data_row(ws, header_row, serial_cols)

    return SheetLayout(
        sheet=ws,
        header_row=header_row,
        data_start_row=data_start,
        groups=groups,
        context=context,
        max_column=ws.max_column,
    )


def detect_workbook(wb) -> SheetLayout:
    """Находит лист с ведомостью СИ в книге. Имя листа не важно."""
    best: SheetLayout | None = None
    for ws in wb.worksheets:
        try:
            layout = detect_layout(ws)
        except Exception as e:
            logger.debug("Лист %r не распознан: %s", ws.title, e)
            continue
        if layout and (best is None or len(layout.groups) > len(best.groups)):
            best = layout

    if best is None:
        raise LayoutNotRecognizedError(
            "Не удалось распознать ведомость СИ. Ожидается таблица с колонками "
            "«Зав. номер» и «Дата поверки» и шапкой с группами приборов "
            "(счётчик, трансформатор тока, трансформатор напряжения)."
        )

    logger.info(
        "Распознан лист %r: шапка в строке %s, данные с %s, группы: %s",
        best.sheet.title, best.header_row, best.data_start_row, list(best.groups),
    )
    return best

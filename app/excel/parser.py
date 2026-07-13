"""Парсер ведомости СИ.

Раскладку определяет detector.py по структуре документа, а не по имени листа
и не по жёстким номерам колонок: метролог волен назвать лист как угодно,
вставить сверху строки с подписями и сдвинуть столбцы.
"""
from __future__ import annotations

import json
import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.excel.detector import (
    LayoutNotRecognizedError,
    SheetLayout,
    detect_workbook,
)
from app.excel.normalize import (
    normalize_serial_for_comparison,
    normalize_serial_for_request,
    normalize_string,
    normalize_type,
    to_canonical_date,
    extract_vri_from_link,
)

logger = logging.getLogger(__name__)

SKIP_MARKERS = {"#name?", "#имя?", "#ref!", "#value!", "#знач!", "#div/0!"}
EMPTY_VALUES = {"", "-", "—", "нет"}


class TemplateNotMatchedError(Exception):
    """Документ не распознан как ведомость СИ."""
    pass


class TemplateDrivenParser:
    """Парсер ведомости СИ со структурным распознаванием раскладки."""

    def __init__(self, template_code: str = "auto", profiles_dir: str = "app/templates_profiles"):
        self.template_code = template_code
        self.profiles_dir = Path(profiles_dir)
        self.layout: SheetLayout | None = None

    def parse_workspace_file(self, file_path: str) -> list[dict]:
        wb = load_workbook(file_path, data_only=True)

        try:
            layout = detect_workbook(wb)
        except LayoutNotRecognizedError as e:
            raise TemplateNotMatchedError(str(e)) from e

        self.layout = layout
        sheet = layout.sheet

        # Колонки для ссылок: если их нет в файле — назначаем новые справа.
        # Аннотатор создаст их и проставит туда ссылки на карточки Аршина.
        link_cols = self._assign_link_columns(layout)

        devices: list[dict] = []
        context: dict[str, str] = {}  # forward-fill: значения тянутся вниз

        for row_idx in range(layout.data_start_row, sheet.max_row + 1):
            row_vals = {
                c: sheet.cell(row=row_idx, column=c).value
                for c in range(1, layout.max_column + 1)
            }
            if all(v is None or str(v).strip() == "" for v in row_vals.values()):
                continue
            if any(
                str(v).strip().lower() in SKIP_MARKERS
                for v in row_vals.values() if v is not None
            ):
                continue

            # Контекст (объект, инвентарный номер и т.п.) заполняется
            # только в первой строке блока — тянем его вниз
            for key, col in layout.context.items():
                val = row_vals.get(col)
                if val is not None and str(val).strip() not in EMPTY_VALUES:
                    context[key] = normalize_string(val)

            for kind, cols in layout.groups.items():
                serial_raw = normalize_serial_for_request(row_vals.get(cols["serial"]))
                if not serial_raw or serial_raw.strip().lower() in EMPTY_VALUES:
                    continue

                type_val = row_vals.get(cols["type"])
                link_col = link_cols[kind]
                link_val = row_vals.get(link_col) if cols.get("arshin_link") else None

                vd_col = cols.get("verification_date")
                nd_col = cols.get("next_verification_date")

                devices.append({
                    "sheet_name": sheet.title,
                    "excel_row": row_idx,
                    "device_kind": kind,
                    "block_code": kind,
                    "type_raw": type_val,
                    "type_norm": normalize_type(type_val),
                    "serial_raw": serial_raw,
                    "serial_norm": normalize_serial_for_comparison(serial_raw),
                    "accuracy_class_raw": row_vals.get(cols.get("accuracy_class", 0)),
                    "verification_date_raw": row_vals.get(vd_col) if vd_col else None,
                    "verification_date_norm": to_canonical_date(row_vals.get(vd_col)) if vd_col else None,
                    "next_date_raw": row_vals.get(nd_col) if nd_col else None,
                    "next_date_norm": to_canonical_date(row_vals.get(nd_col)) if nd_col else None,
                    "link_raw": link_val,
                    "link_vri": extract_vri_from_link(link_val),
                    # Колонки ссылок может не быть в исходном файле — тогда её создадут
                    "link_column_exists": bool(cols.get("arshin_link")),
                    "context": context.copy(),
                    "cell_refs": {
                        "type": self._ref(row_idx, cols["type"]),
                        "serial": self._ref(row_idx, cols["serial"]),
                        "verification_date": self._ref(row_idx, vd_col) if vd_col else None,
                        "next_date": self._ref(row_idx, nd_col) if nd_col else None,
                        "link": self._ref(row_idx, link_col),
                    },
                })

        logger.info(
            "Разобран лист %r: приборов %s (группы: %s)",
            sheet.title, len(devices), list(layout.groups),
        )
        return devices

    def _assign_link_columns(self, layout: SheetLayout) -> dict[str, int]:
        """Колонка ссылки для каждой группы.

        Если в файле её нет, резервируем новые колонки справа от таблицы —
        по одной на группу, в том же порядке, что и группы.
        """
        link_cols: dict[str, int] = {}
        next_free = layout.max_column + 1

        for kind, cols in layout.groups.items():
            existing = cols.get("arshin_link")
            if existing:
                link_cols[kind] = existing
            else:
                link_cols[kind] = next_free
                next_free += 1
                logger.info(
                    "Для группы %s колонка ссылки отсутствует — будет создана: %s",
                    kind, get_column_letter(link_cols[kind]),
                )
        return link_cols

    @staticmethod
    def _ref(row: int, col: int | None) -> str | None:
        if not col:
            return None
        return f"{get_column_letter(col)}{row}"

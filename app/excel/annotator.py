"""Разметка результата в Excel: заливка, комментарии, ссылки на карточки Аршина.

Адреса ячеек приходят готовыми из парсера (cell_refs), поэтому раскладку
здесь заново определять не нужно — достаточно найти тот же лист.
Если колонки для ссылки в исходном файле не было, она создаётся.
"""
from __future__ import annotations

import logging

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from app.models.enums import JobIssueSeverity

logger = logging.getLogger(__name__)

COLOR_MAP = {
    JobIssueSeverity.RED: "FFFFC7CE",
    JobIssueSeverity.YELLOW: "FFFFEB9C",
    JobIssueSeverity.ORANGE: "FFFFCC99",
    JobIssueSeverity.INFO: None,
}

GROUP_TITLES = {
    "si": "Счётчик",
    "ct": "Трансформатор тока",
    "vt": "Трансформатор напряжения",
}


class ExcelAnnotator:
    """Размечает исходный файл, сохраняя его форматирование."""

    def __init__(self, template_code: str = "auto", profiles_dir: str = "app/templates_profiles"):
        self.template_code = template_code

    def annotate_and_save(self, input_path: str, output_path: str, results: list[dict]):
        # data_only=False — чтобы сохранить формулы и стили исходного файла
        wb = load_workbook(input_path, data_only=False)

        sheet_name = next((r.get("sheet_name") for r in results if r.get("sheet_name")), None)
        if not sheet_name or sheet_name not in wb.sheetnames:
            raise ValueError("Лист с ведомостью не найден в файле")
        sheet = wb[sheet_name]

        self._ensure_link_columns(sheet, results)

        for result in results:
            self._annotate_device(sheet, result)

        wb.save(output_path)
        logger.info("Размеченный файл сохранён: %s", output_path)

    def _ensure_link_columns(self, sheet, results: list[dict]):
        """Создаёт колонку «Ссылка на карточку в Аршине», если её не было.

        Метролог мог прислать ведомость без такой колонки — тогда мы её
        добавляем справа и подписываем, чтобы результат был самодостаточным.
        """
        new_cols: dict[int, str] = {}
        header_row = None

        for r in results:
            if r.get("link_column_exists"):
                continue
            ref = (r.get("cell_refs") or {}).get("link")
            if not ref:
                continue
            letter, row = coordinate_from_string(ref)
            col = column_index_from_string(letter)
            new_cols.setdefault(col, r.get("device_kind", ""))
            if header_row is None or row < header_row:
                header_row = row

        if not new_cols or header_row is None:
            return

        # Заголовок ставим на строку выше первой строки данных
        title_row = max(1, header_row - 1)
        for col, kind in sorted(new_cols.items()):
            cell = sheet.cell(row=title_row, column=col)
            if not cell.value:
                title = GROUP_TITLES.get(kind, kind)
                cell.value = f"Ссылка на карточку в Аршине ({title})"
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            sheet.column_dimensions[get_column_letter(col)].width = 42

        logger.info(
            "Созданы колонки для ссылок: %s",
            [get_column_letter(c) for c in sorted(new_cols)],
        )

    def _anchor(self, sheet, cell_ref: str) -> str:
        """Верхняя левая ячейка объединённого диапазона (писать можно только в неё)."""
        letter, row = coordinate_from_string(cell_ref)
        col = column_index_from_string(letter)
        for rng in sheet.merged_cells.ranges:
            if rng.min_col <= col <= rng.max_col and rng.min_row <= row <= rng.max_row:
                return f"{get_column_letter(rng.min_col)}{rng.min_row}"
        return cell_ref

    def _annotate_device(self, sheet, result: dict):
        issues = result.get("issues", [])
        selected_url = result.get("selected_url")

        for issue in issues:
            cell_ref = issue.get("cell_ref")
            if not cell_ref:
                continue
            anchor = self._anchor(sheet, cell_ref)
            cell = sheet[anchor]

            color = COLOR_MAP.get(issue.get("severity"))
            if color:
                cell.fill = PatternFill(fill_type="solid", fgColor=color)

            message = issue.get("message", "")
            if message:
                text = f"{cell.comment.text}\n\n{message}" if cell.comment else message
                cell.comment = self._make_comment(text)

        # Ссылка на карточку прибора в Аршине
        link_ref = (result.get("cell_refs") or {}).get("link")
        if link_ref and selected_url:
            anchor = self._anchor(sheet, link_ref)
            cell = sheet[anchor]
            cell.value = selected_url
            cell.hyperlink = selected_url
            cell.font = Font(color="0563C1", underline="single", size=9)

    def _make_comment(self, text: str) -> Comment:
        """Комментарий с окном под объём текста.

        По умолчанию openpyxl делает крошечный квадратик, в котором длинное
        сообщение обрезается и его приходится растягивать вручную.
        """
        CHARS_PER_LINE = 45
        lines = sum(max(1, -(-len(p) // CHARS_PER_LINE)) for p in text.split("\n"))

        comment = Comment(text, "Проверка по ФГИС «Аршин»")
        comment.width = 380
        comment.height = max(90, min(20 * lines + 30, 600))
        return comment
